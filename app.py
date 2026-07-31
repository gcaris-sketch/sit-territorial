import subprocess
import sys
import os
import time
import urllib.parse
import re
import io
import base64
import difflib
from datetime import datetime, date, timedelta
import unicodedata
import pandas as pd
import geopandas as gpd
import streamlit as st  # 1. IMPORTANTE: Nos aseguramos de importar Streamlit aquí

# --- PEGA LAS IMPORTACIONES AQUÍ ---
from branca.element import MacroElement
from jinja2 import Template

# --- REEMPLAZAR LA CLASE EasyPrint POR ESTA ---
class EasyPrint(MacroElement):
    def __init__(self, filename="Mapa", position="topleft", export_only=True, size_modes=None, **kwargs):
        super().__init__()
        self._name = "EasyPrint"
        self.filename = filename
        self.position = position
        self._template = Template("""
            {% macro header(this, kwargs) %}
                <script src="https://cdn.jsdelivr.net/npm/leaflet-easyprint@2.1.9/dist/bundle.js"></script>
            {% endmacro %}
            {% macro script(this, kwargs) %}
                L.easyPrint({
                    title: 'Descargar Mapa',
                    position: '{{ this.position }}',
                    filename: '{{ this.filename }}',
                    exportOnly: true,
                    hideControlContainer: true
                }).addTo({{ this._parent.get_name() }});
            {% endmacro %}
        """)

def limpiar_comuna(texto):
    """
    Limpia y normaliza el texto de una comuna:
    - Convierte a mayúsculas.
    - Quita espacios al inicio y final.
    - Elimina tildes y caracteres especiales.
    """
    if pd.isna(texto) or texto is None:
        return ""
    
    texto = str(texto).strip().upper()
    texto = ''.join(
        c for c in unicodedata.normalize('NFD', texto)
        if unicodedata.category(c) != 'Mn'
    )
    return texto

# 2. CAMBIO CLAVE: Decorador de caché agregado a la función de procesamiento
@st.cache_data
def mapear_comunas_eficiente(df_ventas, col_comuna_origen, lista_comunas_oficiales):
    """
    Asocia las comunas de las Notas de Venta con la lista oficial a máxima velocidad.
    Gracias a @st.cache_data, esto se calcula 1 SOLA VEZ en la memoria RAM.
    """
    if df_ventas.empty or col_comuna_origen not in df_ventas.columns:
        return df_ventas

    mapa_oficial = {limpiar_comuna(c): c for c in lista_comunas_oficiales if c != "REGION COMPLETA (Ver Todo)"}
    
    uniques_raw = df_ventas[col_comuna_origen].dropna().unique()
    diccionario_coincidencias = {}

    for raw in uniques_raw:
        clean_raw = limpiar_comuna(raw)
        coincidencia = None

        if clean_raw in mapa_oficial:
            coincidencia = mapa_oficial[clean_raw]
        else:
            for clean_oficial, nombre_oficial in mapa_oficial.items():
                if clean_oficial in clean_raw or clean_raw in clean_oficial:
                    coincidencia = nombre_oficial
                    break

        diccionario_coincidencias[raw] = coincidencia if coincidencia else "Sin Coincidencia Territorial"

    df_ventas['ferreteria_comuna'] = df_ventas[col_comuna_origen].map(diccionario_coincidencias).fillna("Sin Coincidencia Territorial")
    
    return df_ventas

# ==============================================================================
# 3. CAMBIO CLAVE: FUNCIONES DE CARGA LIGERAS CON CACHÉ
# ==============================================================================
# Agrega estas funciones aquí y úsalas para leer tus archivos pesados:

@st.cache_data
def cargar_excel_cache(ruta_o_bytes):
    """Lee el archivo Excel 1 sola vez y lo guarda en la RAM."""
    return pd.read_excel(ruta_o_bytes)

@st.cache_data
def cargar_mapa_cache(ruta_o_bytes):
    """Lee el archivo geográfico/mapa 1 sola vez y lo guarda en la RAM."""
    return gpd.read_file(ruta_o_bytes)

# ==============================================================================
# INSTALADOR DE DEPENDENCIAS
# ==============================================================================
try:
    import fiona
    import geopandas as gpd
    import pandas as pd
    import plotly.express as px
    import requests
    from bs4 import BeautifulSoup
    import kaleido
    import matplotlib.pyplot as plt
except ModuleNotFoundError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "fiona", "geopandas", "openpyxl", "plotly", "requests", "beautifulsoup4", "kaleido", "matplotlib", "python-calamine"])
    import fiona
    import geopandas as gpd
    import pandas as pd
    import plotly.express as px
    import requests
    from bs4 import BeautifulSoup
    import kaleido
    import matplotlib.pyplot as plt

import streamlit as st
import folium
from streamlit_folium import st_folium
from folium.plugins import HeatMap, MarkerCluster

fiona.drvsupport.supported_drivers['Shapefile'] = 'r'

st.set_page_config(layout="wide", page_title="SIT - Inteligencia Territorial Avanzada")

# ==============================================================================
# INICIALIZACIÓN GLOBAL DE VARIABLES
# ==============================================================================
df_ferreterias = pd.DataFrame()
df_obras = pd.DataFrame()
df_ventas = pd.DataFrame()
df_clientes = pd.DataFrame()
gdf_mapa = gpd.GeoDataFrame()
col_comuna_shp = "comuna"

# ==============================================================================
# LISTA MAESTRA DE COMUNAS OFICIALES DE O'HIGGINS
# ==============================================================================
COMUNAS_OHIGGINS_OFICIALES = [
    "REGION COMPLETA (Ver Todo)",
    "Chépica", "Chimbarongo", "Codegua", "Coinco", "Coltauco", "Doñihue", "Graneros", 
    "La Estrella", "Las Cabras", "Litueche", "Lolol", "Machalí", "Malloa", "Marchigüe", 
    "Mostazal", "Nancagua", "Navidad", "Olivar", "Palmilla", "Paredones", "Peralillo", 
    "Peumo", "Pichidegua", "Pichilemu", "Placilla", "Pumanque", "Quinta de Tilcoco", 
    "Rancagua", "Rengo", "Requínoa", "San Fernando", "San Vicente", "Santa Cruz"
]

# ==============================================================================
# FUNCIONES DE NORMALIZACIÓN Y CRUCE DE TEXTO (FUZZY MATCHING CRM OPTIMIZADO)
# ==============================================================================
def normalizar_texto(texto):
    if pd.isna(texto):
        return ""
    t = str(texto).lower().strip()
    t = re.sub(r'[áàäâ]', 'a', t)
    t = re.sub(r'[éèëê]', 'e', t)
    t = re.sub(r'[íìïî]', 'i', t)
    t = re.sub(r'[óòöô]', 'o', t)
    t = re.sub(r'[úùüû]', 'u', t)
    t = re.sub(r'\b(spa|ltda|limitada|eirl|s\.a\.|sa|ferreteria|comercial|e|vía|concesionaria)\b', '', t)
    t = re.sub(r'[^a-z0-9 ]', '', t)
    return " ".join(t.split())

# ==============================================================================
# TRADUCTOR DE TIPOLOGÍAS SEIA Y LÓGICA DE COMPATIBILIDAD
# ==============================================================================
TRADUCTOR_SEIA = {
    'a': 'Acueductos y Embalses', 'b': 'Líneas de Transmisión / Energía', 'c': 'Centrales Generadoras de Energía',
    'd': 'Aeropuertos y Terminales', 'e': 'Puertos, Terminales y Vías de Navegación', 'f': 'Astilleros',
    'g': 'Saneamiento Ambiental / Alcantarillado', 'g1': 'Saneamiento Ambiental / Agua Potable',
    'g.1': 'Saneamiento Ambiental / Agua Potable', 'g2': 'Sistemas de Tratamiento de Residuos',
    'h': 'Proyectos Industriales o Inmobiliarios', 'h1': 'Proyectos Industriales / Manufactura',
    'h.1': 'Proyectos Industriales / Manufactura', 'h2': 'Proyectos Inmobiliarios / Habitacionales',
    'h.2': 'Proyectos Inmobiliarios / Habitacionales', 'i': 'Minería y Extracción', 'i1': 'Desarrollo Minero',
    'i.1': 'Desarrollo Minero', 'j': 'Oleoductos y Gasoductos', 'k': 'Instalaciones Nucleares',
    'l': 'Agroindustrias y Criaderos', 'm': 'Desarrollo Forestal / Celulosas', 'n': 'Explotación Acuícola',
    'o': 'Producción / Almacenamiento Químico', 'p': 'Políticas / Planes Urbanos',
    'q': 'Proyectos Turísticos / Áreas Protegidas', 'r': 'Canchas de Golf / Grandes Recintos',
    's': 'Instalaciones Deportivas / Espectáculos'
}

def decodificar_seia(codigo):
    if pd.isna(codigo): return "No Especificado"
    cod_clean = str(codigo).strip().lower().replace('.', '')
    if cod_clean in TRADUCTOR_SEIA: return TRADUCTOR_SEIA[cod_clean]
    letra_base = cod_clean[0] if len(cod_clean) > 0 else ''
    if letra_base in TRADUCTOR_SEIA: return f"{TRADUCTOR_SEIA[letra_base]} ({codigo})"
    return str(codigo).upper()

def calcular_compatibilidad(sector, codigo_proyecto):
    sector_str = str(sector).lower()
    cod_str = str(codigo_proyecto).lower()
    if "inmobiliario" in sector_str or "vivienda" in sector_str or cod_str.startswith('h'):
        return {"nivel": "ALTA", "color": "#27ae60", "desc": "Consumo garantizado de materiales de obra gruesa y ferretería general."}
    elif "transporte" in sector_str or "hidráulica" in sector_str or "infraestructura" in sector_str or cod_str.startswith('a') or cod_str.startswith('d'):
        return {"nivel": "MEDIA", "color": "#e67e22", "desc": "Demanda enfocada en herramientas, fierro, cemento y soluciones viales."}
    else:
        return {"nivel": "BAJA", "color": "#c0392b", "desc": "Infraestructura pesada. Compras de baja escala local."}

def extraer_correo_seia(url):
    if pd.isna(url) or str(url).strip() == "" or not str(url).startswith("http"): return "Enlace inválido."
    try:
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
        response = requests.get(url, headers=headers, timeout=6)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            correos = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,4}', soup.get_text())
            filtrados = [c for c in correos if "no-responder" not in c.lower() and "seia" not in c.lower()]
            return f"Contactos: {', '.join(list(set(filtrados))[:3])}" if filtrados else "No se encontraron correos directos."
        return f"Error (Código: {response.status_code})"
    except Exception as e:
        return f"Error de conexión: {str(e)}"

LAT_MIN, LAT_MAX = -35.25, -33.45
LON_MIN, LON_MAX = -72.20, -69.60

def filtrar_coordenadas_validas(df):
    if df.empty or 'lat' not in df.columns or 'lon' not in df.columns: return df
    return df[(df['lat'] >= LAT_MIN) & (df['lat'] <= LAT_MAX) & (df['lon'] >= LON_MIN) & (df['lon'] <= LON_MAX)]

# ==============================================================================
# ESTILOS CSS PERSONALIZADOS
# ==============================================================================
st.markdown("""
    <style>
        .licencia-header { text-align: right; font-size: 11px; color: #95a5a6; font-style: italic; margin-top: -30px; margin-bottom: 15px; }
        .main-title { font-family: Arial, sans-serif; color: #1a5276; font-size: 30px; font-weight: 700; margin-bottom: 5px; }
        .main-subtitle { font-family: Arial, sans-serif; color: #7f8c8d; font-size: 15px; margin-bottom: 25px; }
        .kpi-container { display: flex; gap: 30px; margin-bottom: 30px; flex-wrap: wrap; }
        .kpi-card { padding-left: 12px; border-left: 4px solid #1a5276; min-width: 150px; margin-bottom: 10px; }
        .kpi-title { font-size: 13px; color: #555555; margin-bottom: 2px; font-weight: 500; }
        .kpi-value { font-size: 24px; font-weight: bold; color: #111111; }
        .list-title-panel { font-size: 18px; font-weight: 700; color: #2c3e50; margin-bottom: 15px; }
        .card-meta { font-size: 12px; color: #7f8c8d; font-family: sans-serif; margin-top: 4px; line-height: 1.6; }
        .monto-style { color: #27ae60; font-weight: bold; font-size: 13px; }
        .section-title { font-size: 20px; font-weight: 700; color: #2c3e50; margin-top: 20px; margin-bottom: 15px; border-left: 4px solid #1a5276; padding-left: 10px; }
        .badge-compatibilidad { display: inline-block; padding: 3px 8px; border-radius: 4px; color: #ffffff; font-size: 11px; font-weight: bold; text-transform: uppercase; margin-top: 5px; }
        .badge-ferreteria { display: inline-block; padding: 2px 6px; border-radius: 4px; background-color: #ebf5fb; color: #2980b9; font-size: 11px; font-weight: 600; margin-bottom: 5px; }
    </style>
""", unsafe_allow_html=True)

st.markdown('<div class="licencia-header">Prototipo desarrollado bajo licencia de Gastón Caris © 2026</div>', unsafe_allow_html=True)
st.markdown('<div class="main-title">Sistema de Inteligencia Comercial Territorial (SIT)</div>', unsafe_allow_html=True)
st.markdown('<div class="main-subtitle">Módulo de Análisis Integrado de Canales e Infraestructura SEIA</div>', unsafe_allow_html=True)

if "comuna_activa" not in st.session_state: 
    st.session_state.comuna_activa = "REGION COMPLETA (Ver Todo)"
if "capas_activas" not in st.session_state: st.session_state.capas_activas = ["Limites Comunales", "Ferreterias (Calor - R: 2km Fijo)", "Calor Obras: Inmobiliario", "Calor Obras: Energía y Minería", "Calor Obras: Otros Sectores"]
if "ruta_ferreterias" not in st.session_state: st.session_state.ruta_ferreterias = []

# ==============================================================================
# CARGA Y EXTRACCIÓN DE DATOS (OPTIMIZADO CON CACHÉ)
# ==============================================================================

import openpyxl
import pandas as pd
import streamlit as st

@st.cache_data
def cargar_excel_con_links(ruta, nombre_columna_link="WEB"):
    # 1. Cargar el DataFrame normal
    df = pd.read_excel(ruta)
    
    try:
        # 2. Cargar el archivo con openpyxl para "escarbar" los links
        wb = openpyxl.load_workbook(ruta)
        ws = wb.active
        
        # 3. Buscar en qué posición está la columna (ej. "WEB")
        col_idx = None
        for cell in ws[1]: # Asume que la fila 1 tiene los títulos
            if cell.value == nombre_columna_link:
                col_idx = cell.column
                break
                
        # 4. Si encuentra la columna, extrae los links reales
        if col_idx is not None:
            enlaces_reales = []
            for row in range(2, len(df) + 2): # Iterar según el tamaño del DataFrame
                celda = ws.cell(row=row, column=col_idx)
                # Si la celda tiene un hipervínculo oculto, lo extrae
                if celda.hyperlink:
                    enlaces_reales.append(celda.hyperlink.target)
                else:
                    enlaces_reales.append(celda.value)
            
            # 5. Sobrescribir la columna en el DataFrame con los links reales
            df[nombre_columna_link] = enlaces_reales
            
    except Exception as e:
        st.error(f"Error al extraer links: {e}")
        
    return df

# === ASÍ ES COMO LO LLAMAS AHORA ===
# Reemplaza tu pd.read_excel actual por esta función:
# Asegúrate de cambiar 'WEB' por el nombre exacto de la columna en tu Excel que tiene la palabra "Ver"

ruta_excel_seia = "data/Proyectos.xlsx" 
df_seia = cargar_excel_con_links(ruta_excel_seia, nombre_columna_link="WEB")
@st.cache_data
def cargar_y_limpiar_ferreterias(ruta):
    if not os.path.exists(ruta): return pd.DataFrame()
    df = pd.read_excel(ruta)
    mapeo_col = {'title': 'nombre', 'address': 'direccion', 'city': 'comuna', 'phone': 'telefono', 'website': 'sitio_web', 'categoryName': 'categoria', 'totalScore': 'calificacion', 'reviewsCount': 'resenas', 'url': 'url_google_maps', 'location/lat': 'lat', 'location/lng': 'lon'}
    df = df.rename(columns=mapeo_col)
    for col in ['nombre', 'direccion', 'comuna', 'telefono', 'sitio_web', 'categoria', 'calificacion', 'resenas', 'url_google_maps', 'lat', 'lon']:
        if col not in df.columns: df[col] = None
    df['lat'] = df['lat'].astype(str).str.replace(',', '.').astype(float)
    df['lon'] = df['lon'].astype(str).str.replace(',', '.').astype(float)
    df = filtrar_coordenadas_validas(df.dropna(subset=['lat', 'lon']))
    df['calificacion'] = df['calificacion'].astype(str).str.replace(',', '.').astype(float)
    df['resenas'] = pd.to_numeric(df['resenas'], errors='coerce').fillna(0).astype(int)
    df['comuna'] = df['comuna'].astype(str).str.strip()
    return df

@st.cache_data
def cargar_y_limpiar_obras(ruta):
    if not os.path.exists(ruta): return pd.DataFrame()
    
    # === AQUÍ LLAMAS A TU NUEVA FUNCIÓN ===
    df = cargar_excel_con_links(ruta, nombre_columna_link="WEB")
    
    df = df.rename(columns={
        'Nombre del Proyecto': 'titulo', 'Comuna': 'comuna', 'Titular': 'empresa', 
        'Tipo de Proyecto': 'tipo_proyecto_codigo', 'Latitud Punto Representativo': 'lat', 
        'Longitud Punto Representativo': 'lon', 'Estado del Proyecto': 'estado_proyecto', 
        'Sector Productivo': 'sector_productivo', 'Tipo de Presentación': 'tipo_presentacion', 
        'WEB': 'url_seia', 'Inversión (MMU$)': 'inversion_raw', 'Fecha Presentación': 'fecha_pres_raw'
    })
    df['lat'] = df['lat'].astype(str).str.replace(',', '.').astype(float)
    df['lon'] = df['lon'].astype(str).str.replace(',', '.').astype(float)
    df = filtrar_coordenadas_validas(df.dropna(subset=['lat', 'lon']))
    df['organismo'] = df['tipo_proyecto_codigo'].apply(decodificar_seia)
    df['comuna'] = df['comuna'].astype(str).str.strip()
    df['monto'] = pd.to_numeric(df['inversion_raw'].astype(str).str.replace(',', '.'), errors='coerce').fillna(0) * 1000000
    df['fecha_inicio'] = pd.to_datetime(df['fecha_pres_raw'], errors='coerce', dayfirst=True)
    df['fecha_fin'] = df['fecha_inicio']
    df = df.dropna(subset=['fecha_inicio'])
    df['anio_inicio'] = df['fecha_inicio'].dt.year
    df = df[df['monto'] > 0]
    return df

ruta_excel_ferreterias = "data/ferreterias_ohiggins_perfecto.xlsx"
df_ferreterias = cargar_y_limpiar_ferreterias(ruta_excel_ferreterias)

ruta_excel_obras = "data/Proyectos.xlsx"
df_obras = cargar_y_limpiar_obras(ruta_excel_obras)

@st.cache_data
def cargar_y_limpiar_dom(ruta):
    if not os.path.exists(ruta): return pd.DataFrame()
    df = pd.read_excel(ruta)
    # Rellenamos nulos para evitar errores
    df = df.fillna("S/D")
    return df

ruta_excel_dom = "data/Permisos_DOM.xlsx"
df_dom = cargar_y_limpiar_dom(ruta_excel_dom)

# ==============================================================================
# MOTOR DE SIMILITUD (FUZZY MATCHING) 
# ==============================================================================
def match_cliente_avanzado(row, df_ferr, df_obr):
    cliente = str(row.get('Cliente_Limpio', '')).upper()
    if not cliente or cliente == 'NAN':
        return {"tipo": "Sin Match", "nombre_asociado": "", "score": 0.0}
    
    mejor_score = 0.0
    mejor_match = ""
    tipo_match = "Sin Match"
    
    # 1. Comparar con Ferreterías
    if not df_ferr.empty:
        for _, f_row in df_ferr.iterrows():
            nombre_ferr = str(f_row.get('nombre', '')).upper()
            score = difflib.SequenceMatcher(None, cliente, nombre_ferr).ratio()
            if score > mejor_score:
                mejor_score = score
                mejor_match = nombre_ferr
                tipo_match = "Ferretería"
                
    # 2. Comparar con Obras (Titular de la empresa)
    if not df_obr.empty:
        for _, o_row in df_obr.iterrows():
            nombre_obra = str(o_row.get('empresa', '')).upper()
            score = difflib.SequenceMatcher(None, cliente, nombre_obra).ratio()
            if score > mejor_score:
                mejor_score = score
                mejor_match = nombre_obra
                tipo_match = "Obra/Titular"
                
    # Umbral mínimo de similitud
    if mejor_score < 0.7:  
        return {"tipo": "Sin Match", "nombre_asociado": "", "score": 0.0}
        
    return {"tipo": tipo_match, "nombre_asociado": mejor_match, "score": mejor_score}


# ==============================================================================
# CARGA DE CLIENTES Y CRUCE INTELIGENTE (OPTIMIZADO Y FLEXIBLE)
# ==============================================================================
@st.cache_data
def procesar_cruce_clientes(ruta, _df_ferr, _df_obr):
    if not os.path.exists(ruta): 
        return pd.DataFrame()
    
    try:
        xls = pd.ExcelFile(ruta)
        
        # --- 1. PROCESAR HOJA 1: ASIGNADOS ---
        df_cartera = pd.read_excel(xls, sheet_name=0) 
        df_cartera['Estado_Cartera'] = 'En Cartera'
        df_cartera.columns = [str(c).strip() for c in df_cartera.columns]
        
        # Mapeo y valores por defecto para Asignados
        df_cartera['Encargado Comercial'] = df_cartera['Vendedor'] if 'Vendedor' in df_cartera.columns else 'SIN ASIGNAR'
        df_cartera['Comuna'] = df_cartera['Comuna'].astype(str).str.strip().str.title() if 'Comuna' in df_cartera.columns else ''
        df_cartera['Comuna'] = df_cartera['Comuna'].replace({'Chƒpica': 'Chépica', 'Nan': '', 'nan': ''})

        # --- 2. PROCESAR HOJA 2: NO ASIGNADOS ---
        if len(xls.sheet_names) > 1:
            df_sin_asignar = pd.read_excel(xls, sheet_name=1)
            df_sin_asignar['Estado_Cartera'] = 'Sin Asignar'
            df_sin_asignar.columns = [str(c).strip() for c in df_sin_asignar.columns]
            
            # En No Asignados el Encargado Comercial es 'SIN ASIGNAR'
            df_sin_asignar['Encargado Comercial'] = 'SIN ASIGNAR'
            df_sin_asignar['Vendedor'] = df_sin_asignar['Vendedor'] if 'Vendedor' in df_sin_asignar.columns else 'SIN VENDEDOR'
            
            # Como No Asignados NO tiene Comuna, asignamos 'Sin Comuna' para evitar quiebres
            df_sin_asignar['Comuna'] = df_sin_asignar['Comuna'].astype(str).str.strip().str.title() if 'Comuna' in df_sin_asignar.columns else 'Sin Comuna'
            
            # Columnas adicionales de Asignados rellenas por defecto
            for col_faltante in ['Tramo sin Atender', 'Tramo en Cartera', 'Cartera']:
                if col_faltante not in df_sin_asignar.columns:
                    df_sin_asignar[col_faltante] = 'Sin Información'
            
            for col_num in ['Q Obra', 'Límite Crédito']:
                if col_num not in df_sin_asignar.columns:
                    df_sin_asignar[col_num] = 0

            df_cli = pd.concat([df_cartera, df_sin_asignar], ignore_index=True)
        else:
            df_cli = df_cartera

        # --- 3. LIMPIEZA GENERAL DE CLIENTE Y VENTA ---
        col_cliente = next((c for c in df_cli.columns if str(c).lower() in ['cliente', 'nombre', 'razon social']), df_cli.columns[0])
        df_cli['Cliente'] = df_cli[col_cliente]
        df_cli['Cliente_Limpio'] = df_cli['Cliente'].astype(str).str.upper().str.strip()
        
        col_venta = next((c for c in df_cli.columns if any(p in str(c).lower() for p in ['venta', 'monto', 'total', 'precio', 'vta'])), None)
        if col_venta:
            venta_temp = pd.to_numeric(df_cli[col_venta], errors='coerce')
            mask_errores = venta_temp.isna()
            if mask_errores.any():
                textos_limpios = df_cli.loc[mask_errores, col_venta].astype(str)
                textos_limpios = textos_limpios.str.replace(r'[^\d,.-]', '', regex=True)
                textos_limpios = textos_limpios.str.replace('.', '', regex=False).str.replace(',', '.', regex=False)
                venta_temp.loc[mask_errores] = pd.to_numeric(textos_limpios, errors='coerce')
            df_cli['Venta'] = venta_temp.fillna(0.0)
        else:
            df_cli['Venta'] = 0.0

        # Match con ferreterías/obras
        df_cli['Match_Tipo'] = "Sin Match"
        df_cli['Match_Asociado'] = ""
        df_cli['Match_Score'] = 0.0
        
        if not _df_ferr.empty or not _df_obr.empty:
            resultados = df_cli.apply(lambda r: match_cliente_avanzado(r, _df_ferr, _df_obr), axis=1)
            df_cli['Match_Tipo'] = resultados.apply(lambda x: x['tipo'])
            df_cli['Match_Asociado'] = resultados.apply(lambda x: x['nombre_asociado'])
            df_cli['Match_Score'] = resultados.apply(lambda x: x['score'])
            
        return df_cli
    except Exception as e:
        st.error(f"Error al leer la planilla de clientes: {e}")
        return pd.DataFrame()

ruta_excel_clientes = "data/CLIENTES RGA JULIO 2026.xlsx"
if not os.path.exists(ruta_excel_clientes):
    st.sidebar.warning("⚠️ No se encontró el archivo Excel de clientes en la ruta especificada.")
df_clientes = procesar_cruce_clientes(ruta_excel_clientes, df_ferreterias, df_obras)

# ==============================================================================
# CARGA DE CARTOGRAFÍA SHP (OPTIMIZADO)
# ==============================================================================
@st.cache_data
def cargar_mapa_base(ruta, lista_oficial):
    if not os.path.exists(ruta): return gpd.GeoDataFrame()
    gdf = gpd.read_file(ruta)
    gdf.columns = gdf.columns.str.lower()
    col = next((c for c in gdf.columns if c in ['comuna', 'nom_com', 'nombre_com', 'nom_comuna']), list(gdf.columns)[0])
    gdf[col] = gdf[col].astype(str).str.strip()
    gdf[col] = gdf[col].replace("Marchihue", "Marchigüe")
    comunas_validas = [c.lower() for c in lista_oficial if c != "REGION COMPLETA (Ver Todo)"]
    return gdf[gdf[col].str.lower().isin(comunas_validas)], col

ruta_recortado = "data/comunas_ohiggins_opt.shp"
gdf_mapa_result = cargar_mapa_base(ruta_recortado, COMUNAS_OHIGGINS_OFICIALES)
if isinstance(gdf_mapa_result, tuple):
    gdf_mapa, col_comuna_shp = gdf_mapa_result

# ==============================================================================
# FILTROS Y CONTROLES SATELLITALES
# ==============================================================================
st.markdown("### Filtros de Control Operativo")
col_panel1, col_panel2, col_panel3, col_panel4 = st.columns([1.2, 1, 1.2, 1])

# 1. Filtro de Territorio / Comuna
with col_panel1:
    idx_comuna = COMUNAS_OHIGGINS_OFICIALES.index(st.session_state.comuna_activa) if st.session_state.comuna_activa in COMUNAS_OHIGGINS_OFICIALES else 0
    st.session_state.comuna_activa = st.selectbox("Filtrar Territorio:", COMUNAS_OHIGGINS_OFICIALES, index=idx_comuna)

if st.session_state.comuna_activa == "REGION COMPLETA (Ver Todo)":
    df_ferr_comuna, df_obras_comuna = df_ferreterias.copy(), df_obras.copy()
    df_clientes_comuna = df_clientes.copy()
else:
    df_ferr_comuna = df_ferreterias[df_ferreterias['comuna'].str.lower().str.contains(st.session_state.comuna_activa.lower(), na=False)] if not df_ferreterias.empty else pd.DataFrame()
    df_obras_comuna = df_obras[df_obras['comuna'].str.lower().str.contains(st.session_state.comuna_activa.lower(), na=False)] if not df_obras.empty else pd.DataFrame()
    df_clientes_comuna = df_clientes[df_clientes['Comuna'].str.lower().str.contains(st.session_state.comuna_activa.lower(), na=False)] if not df_clientes.empty else pd.DataFrame()

# 2. Buscador de Ferreterías
with col_panel2: 
    busqueda_ferr = st.text_input("Buscar Ferretería por Nombre:", placeholder="Nombre...")

# 3. Sector Productivo y Etapa Estimada
with col_panel3:
    lista_sectores = sorted(df_obras_comuna['sector_productivo'].dropna().astype(str).unique().tolist()) if not df_obras_comuna.empty else []
    sectores_seleccionados = st.multiselect("Sector Productivo SEIA:", options=lista_sectores, default=lista_sectores) if lista_sectores else []
    if not lista_sectores: 
        st.text_input("Sector Productivo:", value="Sin obras", disabled=True)
    
    opcion_etapa = st.selectbox(
        "Etapa Estimada del Proyecto:",
        ["Todas las etapas", "Por Empezar (< 6 meses)", "En Construcción (> 6 meses)", "En Calificación / Trámite"]
    )

# 4. Vigencia / Rango de Fechas (En Español)
with col_panel4:
    if not df_obras_comuna.empty and 'fecha_inicio' in df_obras_comuna.columns:
        fecha_max_data = df_obras_comuna['fecha_inicio'].max().date()
        fecha_min_data = df_obras_comuna['fecha_inicio'].min().date()
        
        opcion_rango = st.selectbox(
            "Vigencia / Rango de Fechas:",
            ["Todo el historial", "Última semana", "Último mes", "Últimos 3 meses", "Últimos 6 meses", "Último año"]
        )

        if opcion_rango == "Última semana":
            f_inicio, f_fin = fecha_max_data - timedelta(days=7), fecha_max_data
        elif opcion_rango == "Último mes":
            f_inicio, f_fin = fecha_max_data - timedelta(days=30), fecha_max_data
        elif opcion_rango == "Últimos 3 meses":
            f_inicio, f_fin = fecha_max_data - timedelta(days=90), fecha_max_data
        elif opcion_rango == "Últimos 6 meses":
            f_inicio, f_fin = fecha_max_data - timedelta(days=180), fecha_max_data
        elif opcion_rango == "Último año":
            f_inicio, f_fin = fecha_max_data - timedelta(days=365), fecha_max_data
        else:
            f_inicio, f_fin = fecha_min_data, fecha_max_data
    else:
        f_inicio, f_fin = None, None
        st.selectbox("Vigencia / Rango de Fechas:", ["Sin datos"], disabled=True)

# ==============================================================================
# APLICACIÓN DE FILTROS EN CADENA
# ==============================================================================
df_final_ferr = df_ferr_comuna[df_ferr_comuna['nombre'].str.contains(busqueda_ferr, case=False, na=False)] if (not df_ferr_comuna.empty and busqueda_ferr) else df_ferr_comuna

df_final_obras = df_obras_comuna.copy()
if not df_final_obras.empty:
    # 1. Filtro por Sector Productivo
    if sectores_seleccionados: 
        df_final_obras = df_final_obras[df_final_obras['sector_productivo'].isin(sectores_seleccionados)]
    
    # 2. Filtro por Rango de Fechas
    if f_inicio and f_fin and 'fecha_inicio' in df_final_obras.columns:
        df_final_obras = df_final_obras[
            (df_final_obras['fecha_inicio'].dt.date >= f_inicio) & 
            (df_final_obras['fecha_inicio'].dt.date <= f_fin)
        ]
        
    # 3. Filtro por Etapa Estimada del Proyecto
    hoy = date.today()
    if opcion_etapa == "Por Empezar (< 6 meses)":
        hace_6_meses = hoy - timedelta(days=180)
        df_final_obras = df_final_obras[
            (df_final_obras['estado_proyecto'].astype(str).str.upper() == "APROBADO") & 
            (df_final_obras['fecha_inicio'].dt.date >= hace_6_meses)
        ]
    elif opcion_etapa == "En Construcción (> 6 meses)":
        hace_6_meses = hoy - timedelta(days=180)
        df_final_obras = df_final_obras[
            (df_final_obras['estado_proyecto'].astype(str).str.upper() == "APROBADO") & 
            (df_final_obras['fecha_inicio'].dt.date < hace_6_meses)
        ]
    elif opcion_etapa == "En Calificación / Trámite":
        df_final_obras = df_final_obras[
            df_final_obras['estado_proyecto'].astype(str).str.upper() != "APROBADO"
        ]

# ==============================================================================
# PANEL DE MÉTRICAS (KPI)
# ==============================================================================
total_visible_ferr, total_visible_obras = len(df_final_ferr), len(df_final_obras)
monto_kpi_millones = (df_final_obras['monto'].sum() / 1000000) if not df_final_obras.empty else 0
total_ventas_territoriales = df_clientes_comuna['Venta'].sum() if not df_clientes_comuna.empty else 0.0

st.markdown(f"""
    <div class="kpi-container">
        <div class="kpi-card"><div class="kpi-title">Foco Territorial</div><div class="kpi-value">{st.session_state.comuna_activa.upper()}</div></div>
        <div class="kpi-card"><div class="kpi-title">Ferreterías</div><div class="kpi-value">{total_visible_ferr} locales</div></div>
        <div class="kpi-card"><div class="kpi-title">Proyectos SEIA</div><div class="kpi-value">{total_visible_obras} frentes</div></div>
        <div class="kpi-card"><div class="kpi-title">Inversión</div><div class="kpi-value">$ {monto_kpi_millones:,.2f} MM USD</div></div>
        <div class="kpi-card"><div class="kpi-title">Venta CRM Local</div><div class="kpi-value">$ {total_ventas_territoriales:,.0f} CLP</div></div>
    </div>
""", unsafe_allow_html=True)

# PLANIFICADOR DE RUTA Y CONFIGURACIÓN BASE
with st.sidebar:
    st.markdown("### Estilo del Mapa")
    mapa_base_seleccionado = st.radio(
        "Selecciona el Mapa Base:", 
        ["Claro (Ejecutivo)", "Físico (Topográfico)"], 
        index=0
    )
    st.markdown("---")
    st.markdown("### Planificador de Rutas")
    if st.session_state.ruta_ferreterias:
        st.markdown(f"**Locales en ruta:** {len(st.session_state.ruta_ferreterias)}")
        puntos_ruta = []
        for i, f_id in enumerate(st.session_state.ruta_ferreterias):
            row_f = df_ferreterias.loc[f_id]
            st.caption(f"{i+1}. {row_f['nombre'].upper()}")
            puntos_ruta.append(f"{row_f['lat']},{row_f['lon']}")
        if puntos_ruta:
            url_gmaps = f"https://www.google.com/maps/dir/?api=1&destination={puntos_ruta[-1]}" + (f"&waypoints={urllib.parse.quote('|'.join(puntos_ruta[:-1]))}" if len(puntos_ruta)>1 else "")
            st.link_button("Exportar a Maps", url_gmaps, use_container_width=True)
        if st.button("Limpiar Ruta"): st.session_state.ruta_ferreterias = []; st.rerun()

# Selector de capas limpio en Streamlit
st.session_state.capas_activas = st.multiselect(
    "Capas y Herramientas del Mapa:", 
    [
        "Limites Comunales", 
        "Coroplético: Densidad Ferreterías",
        "Coroplético: Inversión Obras (USD)",
        "Ferreterias (Puntos)", 
        "Ferreterias (Clusters)", 
        "Ferreterias (Calor - R: 2km Fijo)", 
        "Obras (Puntos)", 
        "Obras (Clusters)", 
        "Calor Obras: Inmobiliario",
        "Calor Obras: Energía y Minería",
        "Calor Obras: Otros Sectores"
    ], 
    default=[
        "Limites Comunales", 
        "Ferreterias (Calor - R: 2km Fijo)", 
        "Calor Obras: Inmobiliario", 
        "Calor Obras: Energía y Minería", 
        "Calor Obras: Otros Sectores"
    ]
)

# ==============================================================================
# AUXILIAR: NORMALIZACIÓN ESTÁNDAR PARA UNIÓN DE CAPAS (Mayúsculas/Minúsculas/Acentos)
# ==============================================================================

import numpy as np

# ==============================================================================
# FUNCION AUXILIAR PARA FORMATO DE MONEDA
# ==============================================================================
def formatear_dinero(monto):
    if pd.isna(monto) or monto == 0:
        return "$ 0"
    try:
        return f"$ {int(monto):,}".replace(",", ".")
    except:
        return "$ 0"

import pandas as pd
import numpy as np

# --- LIMPIEZA ROBUSTA DE ENCARGADO COMERCIAL Y VENDEDOR ---
if 'Encargado Comercial' in df_clientes.columns:
    df_clientes['Encargado Comercial'] = df_clientes['Encargado Comercial'].fillna('SIN ASIGNAR').astype(str).str.strip()
if 'Vendedor' in df_clientes.columns:
    df_clientes['Vendedor'] = df_clientes['Vendedor'].fillna('SIN VENDEDOR').astype(str).str.strip()

# 1. Eliminar a Valeria Garrido y Carolina Lopez de TODA la base
nombres_eliminar = ['VALERIA GARRIDO', 'CAROLINA LOPEZ', 'CAROLINA LÓPEZ']
mask_eliminar = df_clientes['Encargado Comercial'].str.upper().isin(nombres_eliminar) | \
                df_clientes['Vendedor'].str.upper().isin(nombres_eliminar)
df_clientes = df_clientes[~mask_eliminar].copy()

# 2. Jonathan Diaz pasa a ser "No Asignado" en todo orden (Encargado o Vendedor)
patron_jonathan = ['JONATHAN DÍAZ', 'JONATHAN DIAZ']
if 'Encargado Comercial' in df_clientes.columns:
    df_clientes.loc[df_clientes['Encargado Comercial'].str.upper().isin(patron_jonathan), 'Encargado Comercial'] = 'SIN ASIGNAR'
if 'Vendedor' in df_clientes.columns:
    df_clientes.loc[df_clientes['Vendedor'].str.upper().isin(patron_jonathan), 'Vendedor'] = 'SIN VENDEDOR'

# 3. Limpieza estándar de nulos y vacíos para Encargado Comercial
if 'Encargado Comercial' in df_clientes.columns:
    patron_null = df_clientes['Encargado Comercial'].str.upper().isin(
        ['NULL', 'NONE', 'NAN', '', 'SIN ASIGNAR', '<NA>', 'N/A']
    ) | df_clientes['Encargado Comercial'].str.lower().str.contains('null', na=False)
    
    df_clientes.loc[patron_null, 'Encargado Comercial'] = 'SIN ASIGNAR'

# --- LIMPIEZA ROBUSTA DE ENCARGADO COMERCIAL Y VENDEDOR ---
if 'Encargado Comercial' in df_clientes.columns:
    df_clientes['Encargado Comercial'] = df_clientes['Encargado Comercial'].fillna('SIN ASIGNAR').astype(str).str.strip()
if 'Vendedor' in df_clientes.columns:
    df_clientes['Vendedor'] = df_clientes['Vendedor'].fillna('SIN VENDEDOR').astype(str).str.strip()

# 1. Eliminar a Valeria Garrido y Carolina Lopez de TODA la base
nombres_eliminar = ['VALERIA GARRIDO', 'CAROLINA LOPEZ', 'CAROLINA LÓPEZ']
mask_eliminar = df_clientes['Encargado Comercial'].str.upper().isin(nombres_eliminar) | \
                df_clientes['Vendedor'].str.upper().isin(nombres_eliminar)
df_clientes = df_clientes[~mask_eliminar].copy()

# 2. Jonathan Diaz pasa a ser "No Asignado"
patron_jonathan = ['JONATHAN DÍAZ', 'JONATHAN DIAZ']
if 'Encargado Comercial' in df_clientes.columns:
    df_clientes.loc[df_clientes['Encargado Comercial'].str.upper().isin(patron_jonathan), 'Encargado Comercial'] = 'SIN ASIGNAR'
if 'Vendedor' in df_clientes.columns:
    df_clientes.loc[df_clientes['Vendedor'].str.upper().isin(patron_jonathan), 'Vendedor'] = 'SIN VENDEDOR'

# 3. Limpieza estándar de nulos y vacíos
if 'Encargado Comercial' in df_clientes.columns:
    patron_null = df_clientes['Encargado Comercial'].str.upper().isin(
        ['NULL', 'NONE', 'NAN', '', 'SIN ASIGNAR', '<NA>', 'N/A', 'S/N', 'SIN ENCARGADO']
    ) | df_clientes['Encargado Comercial'].str.lower().str.contains('null', na=False)
    
    df_clientes.loc[patron_null, 'Encargado Comercial'] = 'SIN ASIGNAR'

# --- SEPARACIÓN DE CARTERAS ---
df_asignados = df_clientes[df_clientes['Estado_Cartera'] == 'En Cartera'].copy()
df_no_asignados = df_clientes[df_clientes['Estado_Cartera'] == 'Sin Asignar'].copy()

# --- 2. FUNCIÓN DE SEMÁFORO PARA TRAMO SIN ATENDER ---
def color_tramo(val):
    """Aplica colores dependiendo de los días sin atender (soporta números o texto)"""
    if pd.isna(val) or val == '':
        return ''
    
    val_str = str(val).lower()
    
    # Intento 1: Si el valor es puramente numérico
    try:
        dias = float(val)
        if dias <= 30: return 'background-color: #c8e6c9; color: #1b5e20;' # Verde
        elif dias <= 90: return 'background-color: #fff9c4; color: #f57f17;' # Amarillo
        else: return 'background-color: #ffcdd2; color: #b71c1c;' # Rojo
    except ValueError:
        pass # Si falla, es porque es texto. Pasamos al intento 2.

    # Intento 2: Si el valor es texto (ej: "1 a 30", "+ de 90")
    if '30' in val_str and ('menor' in val_str or '1 a' in val_str or '0' in val_str):
        return 'background-color: #c8e6c9; color: #1b5e20;'
    elif '90' in val_str and ('mas' in val_str or 'más' in val_str or '>' in val_str or '+' in val_str):
        return 'background-color: #ffcdd2; color: #b71c1c;'
    elif '60' in val_str or '90' in val_str:
        return 'background-color: #fff9c4; color: #f57f17;'
    
    return ''

# ==============================================================================
# CONFIGURACIÓN DE INICIO Y NAVEGACIÓN PRINCIPAL
# ==============================================================================
nav_principal = st.radio(
    "Selecciona la vista que deseas cargar:",
    [
        "🗺️ 1. Mapa Oportunidades y Obras",
        "👥 2. Analisis Cartera de Clientes",
        "🎯 3. Dominio Territorial y Cartera", # <--- CAMBIADO
        "🏗️ 4. Permisos Menores DOM (O'Higgins)"
    ],
    horizontal=True,
    label_visibility="collapsed",
    key="nav_principal_radio"
)

# ==============================================================================
# VISTA 1: MAPA DE OPORTUNIDADES E INFRAESTRUCTURA
# ==============================================================================
if nav_principal == "🗺️ 1. Mapa Oportunidades y Obras":
    col_izq, col_der = st.columns([11, 7])
    
    with col_izq:
        # Configuración del centro del mapa
        if st.session_state.comuna_activa == "REGION COMPLETA (Ver Todo)": 
            centro_lat, centro_lon, zoom_dinamico = -34.4200, -71.0500, 9
        elif not gdf_mapa.empty: 
            gdf_comuna_pol = gdf_mapa[gdf_mapa[col_comuna_shp].str.lower().str.contains(st.session_state.comuna_activa.lower(), na=False)]
            if not gdf_comuna_pol.empty: 
                centro_lat, centro_lon, zoom_dinamico = gdf_comuna_pol.geometry.centroid.y.iloc[0], gdf_comuna_pol.geometry.centroid.x.iloc[0], 11
            else: 
                centro_lat, centro_lon, zoom_dinamico = -34.4200, -71.0500, 9
        else: 
            centro_lat, centro_lon, zoom_dinamico = -34.4200, -71.0500, 9

        tiles_url = "https://{s}.tile.opentopomap.org/{z}/{x}/{y}.png" if mapa_base_seleccionado == "Físico (Topográfico)" else "cartodb positron"
        attribution = "Map data: &copy; OpenStreetMap | Style: &copy; OpenTopoMap" if mapa_base_seleccionado == "Físico (Topográfico)" else "CartoDB"

        m = folium.Map(location=[centro_lat, centro_lon], zoom_start=zoom_dinamico, tiles=tiles_url, attr=attribution)

        # 1. CAPA: LÍMITES COMUNALES
        if "Limites Comunales" in st.session_state.capas_activas and not gdf_mapa.empty:
            for idx, row in gdf_mapa.iterrows():
                nombre_c = row[col_comuna_shp]
                es_activa = (st.session_state.comuna_activa.lower() in nombre_c.lower())
                folium.GeoJson(
                    row.geometry.__geo_interface__, 
                    style_function=lambda x, est={
                        'fillColor': '#1a5276' if es_activa else '#7f8c8d', 
                        'color': '#2c3e50' if es_activa else '#7f8c8d', 
                        'weight': 2.0 if es_activa else 0.8, 
                        'fillOpacity': 0.12 if es_activa else 0.0
                    }: est, 
                    tooltip=f"Comuna: {nombre_c.upper()}"
                ).add_to(m)

        # 2. CAPAS: FERRETERÍAS
        if not df_final_ferr.empty:
            if "Ferreterias (Puntos)" in st.session_state.capas_activas:
                for idx, row in df_final_ferr.iterrows():
                    tel_txt = f"☎ {row.get('telefono')}" if pd.notnull(row.get('telefono')) else "☎ S/N"
                    html = f"<div style='width:220px;'><b>{row['nombre'].upper()}</b><br>⭐ {row.get('calificacion','S/N')}<br>📍 {row['direccion']}<br>{tel_txt}</div>"
                    
                    folium.CircleMarker(
                        location=[row["lat"], row["lon"]], 
                        radius=5, 
                        popup=folium.Popup(html, max_width=250), 
                        color="#1f618d", 
                        weight=0.5, 
                        fill=True, 
                        fill_color="#3498db", 
                        fill_opacity=0.85
                    ).add_to(m)
            
            if "Ferreterias (Calor - R: 2km Fijo)" in st.session_state.capas_activas:
                for _, r in df_final_ferr.iterrows():
                    folium.Circle(location=[r["lat"], r["lon"]], radius=2000, stroke=False, fill=True, fill_color="#2980b9", fill_opacity=0.08).add_to(m)

        # 3. CAPAS: OBRAS
        if not df_final_obras.empty:
            if "Obras (Puntos)" in st.session_state.capas_activas:
                for _, row in df_final_obras.iterrows():
                    comp = calcular_compatibilidad(row.get('sector_productivo', ''), row.get('tipo_proyecto_codigo', ''))
                    html_obra = f"""
                    <div style="width:260px; font-family: Arial, sans-serif; font-size:12px; color:#333;">
                        <h4 style="margin: 0 0 8px 0; color:#2c3e50; border-bottom: 2px solid {comp['color']}; padding-bottom: 4px; font-size:13px;">
                            🏗️ {row['titulo'].upper()}
                        </h4>
                        <p style="margin: 4px 0;"><b>Sector:</b> {row.get('sector_productivo', 'No especificado')}</p>
                        <p style="margin: 4px 0;"><b>Inversión:</b> <span style="color:#27ae60; font-weight:bold;">{formatear_dinero(row['monto'])} USD</span></p>
                        <p style="margin: 4px 0;"><b>Estado:</b> {row.get('estado_proyecto', 'En Calificación')}</p>
                    </div>
                    """
                    folium.CircleMarker(
                        location=[row["lat"], row["lon"]], 
                        radius=5.5, 
                        popup=folium.Popup(html_obra, max_width=280), 
                        tooltip=f"Obra: {row['titulo']}", 
                        color=comp["color"], 
                        weight=0.5, 
                        fill=True, 
                        fill_color=comp["color"], 
                        fill_opacity=0.90
                    ).add_to(m)
            
            comuna_activa_clean = st.session_state.comuna_activa.lower()
            radio_obras_metros = 3000 if any(u in comuna_activa_clean for u in ["rancagua", "machalí", "machali"]) else 5000 
            
            df_obras_heat = df_final_obras.copy()
            df_obras_heat['tipo_clean'] = df_obras_heat['tipo_proyecto_codigo'].astype(str).str.lower().str.strip()
            df_obras_heat['sector_clean'] = df_obras_heat['sector_productivo'].astype(str).str.lower()
            
            df_inmob = df_obras_heat[df_obras_heat['sector_clean'].str.contains('inmobiliario|vivienda', na=False) | df_obras_heat['tipo_clean'].str.startswith('h')]
            df_ener_min = df_obras_heat[df_obras_heat['sector_clean'].str.contains('energía|energia|minería|mineria', na=False) | df_obras_heat['tipo_clean'].str.startswith(('b', 'c', 'i'))]
            df_otros = df_obras_heat[~df_obras_heat.index.isin(df_inmob.index) & ~df_obras_heat.index.isin(df_ener_min.index)]
            
            if "Calor Obras: Inmobiliario" in st.session_state.capas_activas and not df_inmob.empty:
                for _, r in df_inmob.iterrows():
                    folium.Circle(location=[r["lat"], r["lon"]], radius=radio_obras_metros, stroke=False, fill=True, fill_color="#27ae60", fill_opacity=0.08).add_to(m)
            if "Calor Obras: Energía y Minería" in st.session_state.capas_activas and not df_ener_min.empty:
                for _, r in df_ener_min.iterrows():
                    folium.Circle(location=[r["lat"], r["lon"]], radius=radio_obras_metros, stroke=False, fill=True, fill_color="#e67e22", fill_opacity=0.14).add_to(m)
            if "Calor Obras: Otros Sectores" in st.session_state.capas_activas and not df_otros.empty:
                for _, r in df_otros.iterrows():
                    folium.Circle(location=[r["lat"], r["lon"]], radius=radio_obras_metros, stroke=False, fill=True, fill_color="#8e44ad", fill_opacity=0.12).add_to(m)

        st_folium(m, width="100%", height=650, key=f"mapa_principal", returned_objects=[])

    with col_der:
        st.markdown('<div class="list-title-panel">Listados de Control Focalizado</div>', unsafe_allow_html=True)
        vista_detalle = st.radio("Ver detalle de:", ["Canales", "Infraestructura"], horizontal=True, label_visibility="collapsed", key="radio_detalle_listado")
        
        if vista_detalle == "Canales":
            if not df_final_ferr.empty:
                for idx, row in df_final_ferr.iterrows():
                    with st.container(border=True):
                        st.markdown(f'<div class="badge-ferreteria">🛠️ {row.get("categoria", "FERRETERÍA").upper()}</div>', unsafe_allow_html=True)
                        st.markdown(f"**{str(row['nombre']).upper()}**")
                        
                        tel_txt = row.get('telefono')
                        tel_str = f"☎ {tel_txt}" if pd.notnull(tel_txt) and tel_txt != "" else "☎ S/N"
                        
                        st.markdown(f"<div class='card-meta'><b>⭐ REP:</b> {row.get('calificacion','S/N')}<br><b>📍 UBIC:</b> {row['direccion']}<br><b>{tel_str}</b></div>", unsafe_allow_html=True)
                        c1, c2, c3 = st.columns(3)
                        with c1: 
                            st.link_button("🗺️ Maps", row.get('url_google_maps') or f"https://www.google.com/maps/search/?api=1&query={row['lat']},{row['lon']}", use_container_width=True)
                        with c3:
                            if idx not in st.session_state.ruta_ferreterias:
                                if st.button("➕ Ruta", key=f"r_{idx}", use_container_width=True): 
                                    st.session_state.ruta_ferreterias.append(idx)
                                    st.rerun()
                            else: 
                                st.button("✅ Ok", disabled=True, key=f"rok_{idx}", use_container_width=True)
            else:
                st.info("Sin registros comerciales en este perímetro.")
                
        elif vista_detalle == "Infraestructura":
            if not df_final_obras.empty:
                for idx, row in df_final_obras.iterrows():
                    with st.container(border=True):
                        st.markdown(f"**🏗️ {str(row['titulo']).upper()}**")
                        comp = calcular_compatibilidad(row.get('sector_productivo', ''), row.get('tipo_proyecto_codigo', ''))
                        estado_real = str(row.get('estado_proyecto', 'En Calificación')).upper()
                        
                        st.markdown(
                            f"<div class='badge-compatibilidad' style='background-color:{comp['color']};'>COMPATIBILIDAD: {comp['nivel']}</div>"
                            f"<div class='card-meta'>"
                            f"<b>ESTADO:</b> <span style='color:#1a5276; font-weight:bold;'>{estado_real}</span><br>"
                            f"<b>INVERSIÓN:</b> <span class='monto-style'>💵 {formatear_dinero(row['monto'])} USD</span><br>"
                            f"<b>TITULAR:</b> 🏢 {str(row.get('empresa','')).upper()}<br>"
                            f"<b>SECTOR:</b> 🚜 {str(row.get('sector_productivo','')).upper()}"
                            f"</div>", 
                            unsafe_allow_html=True
                        )
                        if pd.notnull(row.get('url_seia')) and str(row.get('url_seia')).startswith("http"):
                            c_lk, c_sc = st.columns(2)
                            with c_lk: 
                                st.link_button("🔗 Ficha SEIA", str(row['url_seia']), use_container_width=True)
                            with c_sc:
                                if st.button("📧 Contacto", key=f"s_{idx}", use_container_width=True): 
                                    st.session_state[f"c_{idx}"] = extraer_correo_seia(row['url_seia'])
                            if f"c_{idx}" in st.session_state: 
                                st.info(st.session_state[f"c_{idx}"])
            else:
                st.info("Sin proyectos de infraestructura registrados.")

    # GRÁFICOS DE RESUMEN AL FINAL DE LA VISTA 1
    st.markdown("---")
    st.markdown("### 📊 Resumen de Infraestructura y Obras Mapeadas")
    c_graf1, c_graf2 = st.columns(2)
    with c_graf1:
        if not df_final_obras.empty:
            fig1 = px.pie(df_final_obras, values='monto', names='sector_productivo', title='Inversión por Sector Productivo (USD)', hole=0.3)
            fig1.update_layout(margin=dict(t=40, b=0, l=0, r=0))
            st.plotly_chart(fig1, use_container_width=True)
        else:
            st.info("Sin datos para graficar inversión.")
            
    with c_graf2:
        if not df_final_obras.empty:
            fig2 = px.histogram(df_final_obras, x='estado_proyecto', title='Cantidad de Obras por Estado', color='estado_proyecto')
            fig2.update_layout(margin=dict(t=40, b=0, l=0, r=0), showlegend=False)
            st.plotly_chart(fig2, use_container_width=True)
        else:
            st.info("Sin datos para graficar estados de obra.")

# ==============================================================================
# VISTA 2: ANÁLISIS CARTERA DE CLIENTES
# ==============================================================================
elif nav_principal == "👥 2. Analisis Cartera de Clientes":
    
    tipo_cliente = st.radio("Selecciona el grupo de clientes a visualizar:", ["Clientes Asignados (Encargado Comercial)", "Clientes No Asignados (Gestión Vendedores)"], horizontal=True, key="radio_tipo_cliente")
    
    if tipo_cliente == "Clientes Asignados (Encargado Comercial)":
        col_m1, col_d1 = st.columns([11, 7])
        
        with col_m1:
            st.markdown("### Concentración Comunitaria (Clientes Asignados)")
            if not df_asignados.empty and not gdf_mapa.empty:
                m_cli = folium.Map(location=[-34.4200, -71.0500], zoom_start=9, tiles="cartodb positron")
                df_c_cli = df_asignados.copy()
                df_c_cli['comuna_match'] = df_c_cli['Comuna'].apply(limpiar_comuna)
                gdf_temp_cli = gdf_mapa.copy()
                gdf_temp_cli['comuna_match'] = gdf_temp_cli[col_comuna_shp].apply(limpiar_comuna)
                df_agg_cli = df_c_cli.groupby('comuna_match').size().reset_index(name='Cantidad_Clientes')
                
                valores_cant = df_agg_cli[df_agg_cli['Cantidad_Clientes'] > 0]['Cantidad_Clientes']
                escala_cuantiles = sorted(list(set(np.percentile(valores_cant, [0, 20, 40, 60, 80, 100])))) if len(valores_cant) >= 4 else None

                folium.Choropleth(
                    geo_data=gdf_temp_cli.__geo_interface__,
                    name="Coroplético Clientes Asignados",
                    data=df_agg_cli,
                    columns=["comuna_match", "Cantidad_Clientes"],
                    key_on="feature.properties.comuna_match",
                    fill_color="YlGnBu",
                    threshold_scale=escala_cuantiles,
                    fill_opacity=0.75,
                    line_color="#ffffff",
                    line_weight=0.8,
                    legend_name="Cantidad de Clientes Asignados",
                    reset=True
                ).add_to(m_cli)
                
                st_folium(m_cli, width="100%", height=480, key="mapa_cli_asig", returned_objects=[])
            elif df_asignados.empty:
                st.warning("⚠️ No hay clientes asignados a un Encargado Comercial (todos figuran como 'SIN ASIGNAR' o no se encontró el archivo Excel).")
            elif gdf_mapa.empty:
                st.info("⚠️ El mapa base (.shp) no está disponible en la ruta especificada, pero puedes revisar el listado a la derecha.")
            else:
                st.info("No hay registros en la cartera asignada para mostrar en la tabla.")

        # NUEVOS GRÁFICOS DE BARRAS APILADAS Y TREEMAP PARA CARTERA ASIGNADA
        st.markdown("---")
        st.markdown("### Análisis de Ventas y Territorio (Cartera Asignada)")
        cg3_asig, cg4_asig = st.columns(2)
        with cg3_asig:
            if not df_asignados.empty:
                fig_asig_bar = px.bar(
                    df_asignados, 
                    x='Encargado Comercial', 
                    y='Venta', 
                    color='Comuna', 
                    title='Ventas por Encargado desglosado por Comuna', 
                    barmode='stack',
                    labels={'Venta': 'Monto de Venta (CLP)'}
                )
                fig_asig_bar.update_layout(margin=dict(t=40, b=0, l=0, r=0))
                st.plotly_chart(fig_asig_bar, use_container_width=True)
        with cg4_asig:
            if not df_asignados.empty:
                df_tree_asig = df_asignados[df_asignados['Venta'] > 0].copy()
                df_tree_asig['Comuna'] = df_tree_asig['Comuna'].replace(['', 'nan', 'Nan'], 'Sin Comuna').fillna('Sin Comuna')
                df_tree_asig['Encargado Comercial'] = df_tree_asig['Encargado Comercial'].replace(['', 'nan', 'Nan'], 'Sin Encargado').fillna('Sin Encargado')
                if not df_tree_asig.empty:
                    fig_asig_tree = px.treemap(
                        df_tree_asig, 
                        path=[px.Constant("Total Ventas Asignadas"), 'Comuna', 'Encargado Comercial'], 
                        values='Venta', 
                        title='Distribución Territorial de Ventas (Comuna > Encargado)'
                    )
                    fig_asig_tree.update_traces(root_color="lightgrey")
                    fig_asig_tree.update_layout(margin=dict(t=40, b=0, l=0, r=0))
                    st.plotly_chart(fig_asig_tree, use_container_width=True)
            else:
                st.info("Sin datos para el gráfico de ventas.")

        with col_d1:
            st.markdown("### 📋 Listado - Cartera Asignada")
            if not df_asignados.empty:
                df_disp_asig = df_asignados.copy()
                df_disp_asig['Venta_Formateada'] = df_disp_asig['Venta'].apply(formatear_dinero)
                
                # Agregamos 'Tramo sin Atender' a las columnas visibles
                cols_v = [c for c in ['Cliente', 'Comuna', 'Encargado Comercial', 'Tramo sin Atender', 'Venta_Formateada'] if c in df_disp_asig.columns]
                
                # Aplicamos el estilo a la tabla
                df_styled = df_disp_asig[cols_v].rename(columns={'Venta_Formateada': 'Monto Venta'})
                if 'Tramo sin Atender' in df_styled.columns:
                    st.dataframe(df_styled.style.map(color_tramo, subset=['Tramo sin Atender']), height=420, use_container_width=True)
                else:
                    st.dataframe(df_styled, height=420, use_container_width=True)

        # GRÁFICOS DE RESUMEN: CARTERA ASIGNADA
        st.markdown("---")
        st.markdown("###  Resumen Gráfico: Cartera Asignada")
        cg1, cg2 = st.columns(2)
        with cg1:
            if not df_asignados.empty:
                fig_a1 = px.pie(df_asignados, names='Encargado Comercial', title='Distribución de Clientes por Encargado', hole=0.4)
                fig_a1.update_layout(margin=dict(t=40, b=0, l=0, r=0))
                st.plotly_chart(fig_a1, use_container_width=True)
        with cg2:
            if not df_asignados.empty:
                fig_a2 = px.pie(df_asignados, values='Venta', names='Encargado Comercial', title='Monto de Ventas (CLP) por Encargado', hole=0.4)
                fig_a2.update_traces(textposition='inside', textinfo='percent+label')
                fig_a2.update_layout(margin=dict(t=40, b=0, l=0, r=0))
                st.plotly_chart(fig_a2, use_container_width=True)

    elif tipo_cliente == "Clientes No Asignados (Gestión Vendedores)":
        st.markdown("### 📋 Cartera de Clientes No Asignados")
        
        if not df_no_asignados.empty:
            df_disp_na = df_no_asignados.copy()
            
            # Columnas solicitadas del Excel
            cols_excel = ['Sucursal', 'Cod. Cliente', 'Cliente', 'Vendedor', 'Sin Asignar', 'Venta', 'LC Disponible']
            
            # Validar que existan todas las columnas en el dataframe
            for col in cols_excel:
                if col not in df_disp_na.columns:
                    df_disp_na[col] = "S/D"
            
            # Preparar copias para cálculo y formateo de montos
            df_disp_na['Venta_Num'] = pd.to_numeric(df_disp_na['Venta'], errors='coerce').fillna(0)
            df_disp_na['LC_Num'] = pd.to_numeric(df_disp_na['LC Disponible'], errors='coerce').fillna(0)
            
            # KPIs de resumen rápido
            k1, k2, k3 = st.columns(3)
            with k1:
                st.metric("Total Clientes No Asignados", f"{len(df_disp_na)} registros")
            with k2:
                st.metric("Total Venta No Asignada", formatear_dinero(df_disp_na['Venta_Num'].sum()))
            with k3:
                st.metric("Total LC Disponible", formatear_dinero(df_disp_na['LC_Num'].sum()))
                
            st.markdown("---")
            
            # Formatear la tabla con la información exacta del Excel
            df_tabla = df_disp_na[cols_excel].copy()
            df_tabla['Venta'] = df_disp_na['Venta_Num'].apply(formatear_dinero)
            df_tabla['LC Disponible'] = df_disp_na['LC_Num'].apply(formatear_dinero)
            
            st.dataframe(df_tabla, height=400, use_container_width=True)

            # GRÁFICOS BASADOS EXCLUSIVAMENTE EN DATOS DEL EXCEL
            st.markdown("---")
            st.markdown("### 📊 Gráficos de Gestión No Asignados")
            
            cg1, cg2 = st.columns(2)
            
            with cg1:
                # Gráfico 1: Ventas por Vendedor
                df_vendedor = df_disp_na.groupby('Vendedor')['Venta_Num'].sum().reset_index()
                fig_vendedor = px.bar(
                    df_vendedor, 
                    x='Vendedor', 
                    y='Venta_Num', 
                    title='Ventas por Vendedor (CLP)',
                    labels={'Venta_Num': 'Monto Venta', 'Vendedor': 'Vendedor'},
                    color='Vendedor'
                )
                fig_vendedor.update_layout(margin=dict(t=40, b=0, l=0, r=0), showlegend=False)
                st.plotly_chart(fig_vendedor, use_container_width=True)
                
            with cg2:
                # Gráfico 2: Ventas y LC Disponible por Sucursal
                df_sucursal = df_disp_na.groupby('Sucursal')[['Venta_Num', 'LC_Num']].sum().reset_index()
                fig_sucursal = px.bar(
                    df_sucursal, 
                    x='Sucursal', 
                    y=['Venta_Num', 'LC_Num'], 
                    title='Venta vs LC Disponible por Sucursal',
                    barmode='group',
                    labels={'value': 'Monto (CLP)', 'variable': 'Métrica'}
                )
                fig_sucursal.update_layout(margin=dict(t=40, b=0, l=0, r=0))
                st.plotly_chart(fig_sucursal, use_container_width=True)
                
        else:
            st.info("No hay registros en la cartera de clientes no asignados.")

# ==============================================================================
# VISTA 3: DOMINIO TERRITORIAL Y CARTERA
# ==============================================================================
elif nav_principal == "🎯 3. Dominio Territorial y Cartera":
    st.markdown("### 🗺️ Panel de Dominio Comercial y Riesgo de Cartera")
    
    if df_clientes.empty:
        st.warning("⚠️ No hay datos de clientes cargados en el sistema.")
    else:
        # Extraer números de los tramos para poder graficarlos
        df_clientes['dias_sin_atender_num'] = pd.to_numeric(df_clientes['Tramo sin Atender'].astype(str).str.extract(r'(\d+)')[0], errors='coerce').fillna(0)
        df_clientes['dias_en_cartera_num'] = pd.to_numeric(df_clientes['Tramo en Cartera'].astype(str).str.extract(r'(\d+)')[0], errors='coerce').fillna(0)
        df_clientes['comuna_match'] = df_clientes['Comuna'].apply(limpiar_comuna)

        # Crear sub-pestañas eliminando la redundancia del mapa base
        tab_asignacion, tab_dominio, tab_riesgo = st.tabs([
            "⚖️ Propuesta de Asignación Oficial", 
            "👑 Dominio por Vendedor", 
            "🚨 Mapa de Clientes Sin Atender"
        ])
        
        # ---------------------------------------------------------
        # PESTAÑA 1: RECOMENDACIÓN DE ASIGNACIÓN OFICIAL SEGÚN VENTA REAL
        # ---------------------------------------------------------
        with tab_asignacion:
            st.markdown("#### Sugerencia de Asignación Oficial basada en Venta Acumulada")
            st.caption("Identifica clientes no asignados formalmente o descalzados, indicando a qué vendedor deberían reasignarse según la venta real.")
            
            # Filtramos clientes con ventas donde el Encargado sea 'SIN ASIGNAR' o distinto al Vendedor
            df_sug = df_clientes[(df_clientes['Venta'] > 0)].copy()
            df_sug['Estado_Asignacion'] = np.where(
                df_sug['Encargado Comercial'] == 'SIN ASIGNAR', 
                'Sin Encargado Oficial', 
                'Asignado'
            )
            
            if not df_sug.empty:
                # Gráfico de barras apiladas por Vendedor y estado de asignación
                fig_propuesta = px.bar(
                    df_sug,
                    x='Vendedor',
                    y='Venta',
                    color='Estado_Asignacion',
                    hover_data=['Cliente', 'Comuna', 'Encargado Comercial'],
                    title='Volumen de Venta Real por Vendedor (Clientes Asignados vs Sin Asignar)',
                    labels={'Venta': 'Venta Acumulada (CLP)', 'Vendedor': 'Vendedor Operativo'},
                    color_discrete_map={'Sin Encargado Oficial': '#e74c3c', 'Asignado': '#2ecc71'},
                    barmode='stack'
                )
                fig_propuesta.update_layout(height=420)
                st.plotly_chart(fig_propuesta, use_container_width=True)
                
                # Tabla rápida con los Top Clientes que requieren asignación urgente
                st.markdown("**Top Clientes Sin Encargado Oficial recomendados para asignación:**")
                top_sin_asig = df_sug[df_sug['Estado_Asignacion'] == 'Sin Encargado Oficial'].sort_values('Venta', ascending=False).head(10)
                if not top_sin_asig.empty:
                    st.dataframe(
                        top_sin_asig[['Cliente', 'Comuna', 'Vendedor', 'Venta']].assign(
                            Venta=top_sin_asig['Venta'].apply(formatear_dinero)
                        ),
                        use_container_width=True,
                        height=250
                    )
            else:
                st.info("No hay datos de ventas disponibles para calcular propuestas de asignación.")

        # ---------------------------------------------------------
        # PESTAÑA 2: VENDEDOR DOMINANTE POR MONTO
        # ---------------------------------------------------------
        with tab_dominio:
            st.markdown("#### Vendedor con Mayor Monto de Venta por Comuna")
            if not gdf_mapa.empty:
                # Calcular ganador por comuna
                df_ventas = df_clientes.groupby(['comuna_match', 'Vendedor'])['Venta'].sum().reset_index()
                idx_ganadores = df_ventas.groupby('comuna_match')['Venta'].idxmax()
                df_ganadores = df_ventas.loc[idx_ganadores]
                
                dict_vend = df_ganadores.set_index('comuna_match')['Vendedor'].to_dict()
                dict_venta = df_ganadores.set_index('comuna_match')['Venta'].to_dict()
                
                vendedores_unicos = df_ganadores['Vendedor'].unique()
                colores_v = ['#e6194B', '#3cb44b', '#4363d8', '#f58231', '#911eb4', '#42d4f4', '#f032e6', '#bfef45', '#fabed4', '#469990']
                mapa_colores = {v: colores_v[i % len(colores_v)] for i, v in enumerate(vendedores_unicos)}

                m_dominio = folium.Map(location=[-34.4200, -71.0500], zoom_start=9, tiles="cartodb positron")
                gdf_temp_dom = gdf_mapa.copy()
                gdf_temp_dom['comuna_match'] = gdf_temp_dom[col_comuna_shp].apply(limpiar_comuna)

               # Capa GeoJson coloreada por vendedor con bordes suavizados
                folium.GeoJson(
                    gdf_temp_dom,
                    style_function=lambda feature: {
                        'fillColor': mapa_colores.get(dict_vend.get(feature['properties']['comuna_match']), '#cccccc'),
                        'color': '#ffffff',      # <--- CAMBIO: Bordes blancos sutiles
                        'weight': 0.5,          # <--- CAMBIO: Grosor muy fino
                        'fillOpacity': 0.65 if feature['properties']['comuna_match'] in dict_vend else 0.05
                    },
                    tooltip=folium.GeoJsonTooltip(fields=[col_comuna_shp], aliases=['Comuna:'])
                ).add_to(m_dominio)

                # Leyenda estática arriba del mapa
                st.write("**Leyenda de Dominio:**")
                leyenda_html = " ".join([f"<span style='background-color:{c}; color:white; padding:3px 8px; border-radius:4px; font-size:12px; margin-right:5px;'>{v}</span>" for v, c in mapa_colores.items()])
                st.markdown(leyenda_html, unsafe_allow_html=True)
                
                st_folium(m_dominio, width="100%", height=480, key="mapa_dominio", returned_objects=[])

        # ---------------------------------------------------------
        # PESTAÑA 3: ZONAS DESATENDIDAS (RIESGO)
        # ---------------------------------------------------------
        with tab_riesgo:
            st.markdown("#### Comunas que concentran mayor tramo sin atender")
            c_m, c_l = st.columns([1, 1])
            
            # Filtramos clientes problemáticos (>30 días sin atender, configurable)
            df_riesgo = df_clientes[df_clientes['dias_sin_atender_num'] >= 30].copy()
            
            with c_m:
                if not gdf_mapa.empty and not df_riesgo.empty:
                    m_riesgo = folium.Map(location=[-34.4200, -71.0500], zoom_start=9, tiles="cartodb positron")
                    df_agg_riesgo = df_riesgo.groupby('comuna_match').size().reset_index(name='Clientes_En_Riesgo')
                    gdf_temp_riesgo = gdf_mapa.copy()
                    gdf_temp_riesgo['comuna_match'] = gdf_temp_riesgo[col_comuna_shp].apply(limpiar_comuna)
                    
                    folium.Choropleth(
                        geo_data=gdf_temp_riesgo.__geo_interface__,
                        data=df_agg_riesgo,
                        columns=["comuna_match", "Clientes_En_Riesgo"],
                        key_on="feature.properties.comuna_match",
                        fill_color="Reds",
                        fill_opacity=0.8,
                        line_color="#ffffff",
                        legend_name="Cantidad de Clientes sin atender"
                    ).add_to(m_riesgo)
                    st_folium(m_riesgo, width="100%", height=400, key="mapa_riesgo", returned_objects=[])
                else:
                    st.info("Sin datos de riesgo o mapa base.")
                    
            with c_l:
                st.markdown("**Clientes críticos (Tramo sin atender > 30)**")
                if not df_riesgo.empty:
                    st.dataframe(df_riesgo[['Cliente', 'Comuna', 'Vendedor', 'Tramo sin Atender', 'Venta']].sort_values('Venta', ascending=False), height=400, use_container_width=True)

        # ---------------------------------------------------------
        # GRÁFICO INFERIOR Y TABLA MAESTRA
        # ---------------------------------------------------------
        # Gráfico por Rangos Categóricos (Estilo Excel)
        st.markdown("#### 📊 Distribución de Ventas por Tramo de Atención y Cartera")
        
        df_rangos = df_clientes[df_clientes['Venta'] > 0].copy()
        
        if not df_rangos.empty:
            fig_rangos = px.bar(
                df_rangos,
                x='Tramo sin Atender',
                y='Venta',
                color='Tramo en Cartera',
                barmode='group',
                hover_data=['Cliente', 'Comuna', 'Vendedor'],
                title='Ventas Acumuladas por Rango: Días Sin Atender vs Días en Cartera',
                labels={
                    'Tramo sin Atender': 'Rango Días Sin Atender (Excel)',
                    'Venta': 'Venta Acumulada (CLP)',
                    'Tramo en Cartera': 'Tramo en Cartera'
                },
                color_discrete_sequence=px.colors.qualitative.Set2
            )
            fig_rangos.update_layout(height=480, xaxis_title="Tramo sin Atender", yaxis_title="Monto Venta (CLP)")
            st.plotly_chart(fig_rangos, use_container_width=True)
        else:
            st.info("Sin registros de ventas para mostrar la distribución por rangos.")

        st.markdown("#### 📋 Listado Detallado de Clientes Maestros")
        columnas_pedidas = [
            'Sucursal', 'Cod. Cliente', 'Cliente', 'Comuna', 'Vendedor',
            'Encargado Comercial', 'Tramo sin Atender', 'Tramo en Cartera', 
            'Q Obra', 'Cartera', 'Venta', 'Límite Crédito', 'LC Disponible'
        ]
        
        df_mostrar = df_clientes.copy()
        for col in columnas_pedidas:
            if col not in df_mostrar.columns:
                df_mostrar[col] = "S/D"
                
        for col in ['Venta', 'Límite Crédito', 'LC Disponible']:
            if col in df_mostrar.columns:
                df_mostrar[col] = pd.to_numeric(df_mostrar[col], errors='coerce').fillna(0)
                df_mostrar[col] = df_mostrar[col].apply(formatear_dinero)

        st.dataframe(
            df_mostrar[columnas_pedidas].style.map(color_tramo, subset=['Tramo sin Atender']), 
            height=420, 
            use_container_width=True
        )

# ==============================================================================
# VISTA 4: PERMISOS MENORES DOM
# ==============================================================================
elif nav_principal == "🏗️ 4. Permisos Menores DOM (O'Higgins)":
    st.markdown("### 🏗️ Monitor de Permisos de Edificación (DOM)")
    st.info("📌 **Nota del Sistema:** Este módulo está diseñado para centralizar y monitorear los permisos de edificación de escala menor para **toda la Región de O'Higgins**. Actualmente, la base de datos se encuentra en fase inicial operando únicamente con registros de la comuna de **San Fernando**.")

    if df_dom.empty:
        st.warning("No se encontró la base de datos de permisos DOM.")
    else:
        # --- FILTRO CLAVE: Omitir Demoliciones ---
        df_dom_filtrado = df_dom[~df_dom['Tipo_Permiso'].astype(str).str.contains('demolici', case=False, na=False, regex=True)].copy()
        
        c_dom_1, c_dom_2 = st.columns([1, 1])
        
        with c_dom_1:
            st.markdown("#### 📊 Distribución por Destino de Uso")
            if not df_dom_filtrado.empty:
                fig_dom_1 = px.pie(df_dom_filtrado, names='Destino_Uso', hole=0.4, title='Proyectos por Destino (Excluye Demolición)')
                fig_dom_1.update_layout(margin=dict(t=30, b=0, l=0, r=0))
                st.plotly_chart(fig_dom_1, use_container_width=True)
            else:
                st.info("No hay datos tras filtrar las demoliciones.")
            
        with c_dom_2:
            st.markdown("#### 📋 Listado de Permisos y Ubicación")
            for idx, row in df_dom_filtrado.iterrows():
                with st.container(border=True):
                    st.markdown(f"**N° {row['Nro_Permiso']} - {row['Tipo_Permiso']}**")
                    st.markdown(
                        f"<div class='card-meta'>"
                        f"<b>PROPIETARIO:</b> {row['Propietario_Solicitante']}<br>"
                        f"<b>DIRECCIÓN:</b> {row['Direccion_Ubicacion']}, {row['Municipalidad']}<br>"
                        f"<b>DESTINO:</b> {row['Destino_Uso']}"
                        f"</div>", 
                        unsafe_allow_html=True
                    )
                    
                    direccion_limpia = str(row['Direccion_Ubicacion']).replace(" ", "+")
                    comuna_limpia = str(row['Municipalidad']).replace(" ", "+")
                    link_maps = f"https://www.google.com/maps/search/?api=1&query={direccion_limpia},+{comuna_limpia},+Chile"
                    st.link_button("📍 Ver en Google Maps", link_maps)